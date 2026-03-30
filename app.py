
import math
import io
import pandas as pd
from openpyxl import load_workbook
import streamlit as st

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Dew & Bubble Point Calculator",
    page_icon="🔬",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Custom CSS ─────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;600&display=swap');

html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif; }

.stApp { background-color: #0f1117; color: #e0e0e0; }

section[data-testid="stSidebar"] {
    background-color: #161b27;
    border-right: 1px solid #2a3040;
}
section[data-testid="stSidebar"] .stMarkdown p,
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] .stNumberInput label,
section[data-testid="stSidebar"] .stSelectbox label {
    color: #a0aec0 !important;
    font-size: 0.82rem;
    letter-spacing: 0.05em;
    text-transform: uppercase;
}

h1 {
    font-family: 'IBM Plex Mono', monospace !important;
    color: #38bdf8 !important;
    font-size: 1.6rem !important;
    letter-spacing: -0.02em;
    border-bottom: 1px solid #2a3040;
    padding-bottom: 0.5rem;
    margin-bottom: 0.2rem !important;
}
h2, h3 {
    font-family: 'IBM Plex Mono', monospace !important;
    color: #94a3b8 !important;
    font-size: 0.9rem !important;
    letter-spacing: 0.08em;
    text-transform: uppercase;
}

[data-testid="metric-container"] {
    background: #161b27;
    border: 1px solid #2a3040;
    border-radius: 6px;
    padding: 1rem 1.2rem;
}
[data-testid="metric-container"] [data-testid="stMetricLabel"] {
    color: #64748b !important;
    font-size: 0.72rem !important;
    letter-spacing: 0.1em;
    text-transform: uppercase;
    font-family: 'IBM Plex Mono', monospace !important;
}
[data-testid="metric-container"] [data-testid="stMetricValue"] {
    color: #38bdf8 !important;
    font-family: 'IBM Plex Mono', monospace !important;
    font-size: 1.6rem !important;
}

[data-testid="stDataFrame"] {
    border: 1px solid #2a3040 !important;
    border-radius: 6px;
}
[data-testid="stDataEditor"] {
    border: 1px solid #2a3040 !important;
    border-radius: 6px !important;
}

.stAlert { border-radius: 4px; font-size: 0.85rem; }

input[type="number"] {
    background-color: #1e2535 !important;
    color: #e2e8f0 !important;
    border: 1px solid #2a3040 !important;
    border-radius: 4px;
    font-family: 'IBM Plex Mono', monospace !important;
}

[data-testid="stFileUploader"] {
    background: #161b27;
    border: 1px dashed #2a3040;
    border-radius: 6px;
    padding: 0.5rem;
}

.stButton button {
    background: #0ea5e9 !important;
    color: #0f1117 !important;
    border: none !important;
    border-radius: 4px !important;
    font-family: 'IBM Plex Mono', monospace !important;
    font-weight: 600 !important;
    letter-spacing: 0.05em;
    padding: 0.5rem 1.5rem !important;
}
.stButton button:hover { background: #38bdf8 !important; }

.result-card {
    background: #161b27;
    border: 1px solid #2a3040;
    border-left: 3px solid #0ea5e9;
    border-radius: 6px;
    padding: 1rem 1.2rem;
    margin-bottom: 0.6rem;
}
.result-card.error { border-left-color: #ef4444; }

hr { border-color: #2a3040 !important; }

details {
    background: #161b27 !important;
    border: 1px solid #2a3040 !important;
    border-radius: 6px !important;
}
summary { color: #94a3b8 !important; font-size: 0.82rem !important; }

#MainMenu {visibility: hidden;}
footer    {visibility: hidden;}
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  CORE CALCULATION LOGIC
# ══════════════════════════════════════════════════════════════════════════════

MOL_UNIT_KEYWORDS = ["MOL%", "MOL %", "MOLE%", "MOLE %"]

def is_mol_unit(unit_str):
    if not unit_str:
        return False
    u = str(unit_str).strip().upper()
    if "PPM" in u:
        return False
    return any(kw in u for kw in MOL_UNIT_KEYWORDS)


def convert_pressure(P_gauge_kgcm2):
    return (P_gauge_kgcm2 + 1.0332) / 1.0332


def wilson_k(Pc_atm, P_atm, omega, Tc_K, T_K):
    return (Pc_atm / P_atm) * math.exp(5.37 * (1.0 + omega) * (1.0 - Tc_K / T_K))


@st.cache_data(show_spinner=False)
def load_database():
    df = pd.read_excel("database.xlsx", header=None, skiprows=7)
    df.columns = [
        'NUMBER', 'COMPONENT', 'MOLE_WT', 'TFP', 'TB',
        'TC', 'PC', 'VC', 'ZC', 'OMEGA',
        'LIQDEN', 'TDEN', 'DIPM',
        'CP_A', 'CP_B', 'CP_C', 'CP_D',
        'LV_B', 'LV_C',
        'DELHG', 'DELGF',
        'ANT_A', 'ANT_B', 'ANT_C', 'TMX', 'TMN',
        'HAR_A', 'HAR_B', 'HAR_C', 'HAR_D', 'HV',
    ]
    df = df.dropna(subset=['COMPONENT'])
    df['COMPONENT'] = df['COMPONENT'].astype(str).str.strip().str.upper()
    return df


def read_composition_sheet(ws):
    mixture = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        name  = row[0]
        value = row[1] if len(row) > 1 else None
        unit  = row[2] if len(row) > 2 else None
        if name is None or value is None:
            continue
        try:
            value = float(value)
        except (TypeError, ValueError):
            continue
        if value == 0:
            continue
        if not is_mol_unit(unit):
            continue
        mixture.append((str(name).strip().upper(), value))
    return mixture


def calculate_dew_bubble(P_atm, component_data, mole_fractions):
    TOL        = 1e-300
    T_n        = 0.25
    step       = 0.25
    iterations = 2500
    best_dew    = (float('inf'), None)
    best_bubble = (float('inf'), None)

    for _ in range(iterations):
        dew_sum = bubble_sum = 0.0
        for z_i, props in zip(mole_fractions, component_data):
            k = wilson_k(props['PC'], P_atm, props['OMEGA'], props['TC'], T_n)
            if k >= TOL:
                dew_sum += z_i / k
            if k != 0:
                bubble_sum += z_i * k
        d = abs(dew_sum    - 1)
        b = abs(bubble_sum - 1)
        if d < best_dew[0]:
            best_dew = (d, T_n)
        if b < best_bubble[0]:
            best_bubble = (b, T_n)
        T_n += step

    return best_dew[1], best_bubble[1]


def process_mixture(mixture, db, P_atm):
    """mixture : list of (component_name, mole_percent)"""
    if not mixture:
        return {"error": "No MOL% components found (all PPM or all zero)"}

    component_data, mole_percents, not_found = [], [], []

    for name, pct in mixture:
        row = db[db['COMPONENT'] == name.strip().upper()]
        if row.empty:
            partial = db[db['COMPONENT'].str.contains(
                name.strip().upper(), na=False, regex=False)]
            not_found.append((name, partial['COMPONENT'].tolist()[:3]))
            continue
        component_data.append(row.iloc[0])
        mole_percents.append(pct)

    total = sum(mole_percents)
    if total == 0:
        return {"error": "All mole percents are zero"}

    mole_fractions = [p / total for p in mole_percents]
    dew_K, bubble_K = calculate_dew_bubble(P_atm, component_data, mole_fractions)

    return {
        "n":          len(component_data),
        "components": [(c['COMPONENT'], round(f * 100, 4))
                       for c, f in zip(component_data, mole_fractions)],
        "dew_K":      dew_K,
        "dew_C":      dew_K - 273.15,
        "bubble_K":   bubble_K,
        "bubble_C":   bubble_K - 273.15,
        "not_found":  not_found,
    }


def run_all_sheets(comp_bytes, db, P_atm):
    wb      = load_workbook(io.BytesIO(comp_bytes), read_only=True, data_only=True)
    raw     = {}
    results = {}
    for name in wb.sheetnames:
        mixture       = read_composition_sheet(wb[name])
        raw[name]     = mixture
        results[name] = process_mixture(mixture, db, P_atm)
    wb.close()
    return results, raw


# ══════════════════════════════════════════════════════════════════════════════
#  SESSION STATE
# ══════════════════════════════════════════════════════════════════════════════
for key, default in {
    "results":      None,
    "raw_mixtures": None,
}.items():
    if key not in st.session_state:
        st.session_state[key] = default


# ══════════════════════════════════════════════════════════════════════════════
#  SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("# ⚙️ Configuration")
    st.markdown("---")
    st.markdown("**Operating Pressure**")
    pressure = st.number_input(
        "Pressure (kg/cm² gauge)",
        min_value=0.0,
        max_value=500.0,
        value=1.5,
        step=0.1,
        format="%.2f",
    )
    P_atm = convert_pressure(pressure)
    st.caption(f"= {P_atm:.4f} ATM absolute")

    st.markdown("---")
    st.markdown("**Method**")
    st.caption("Wilson Correlation (K-value)\nT range: 0.25 – 625 K, step 0.25 K")


# ══════════════════════════════════════════════════════════════════════════════
#  HEADER
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("# 🔬 Dew & Bubble Point Calculator")
st.markdown(
    "<p style='color:#64748b;font-size:0.85rem;margin-top:-0.5rem'>"
    "Multi-sheet composition table · MOL% components only"
    "</p>",
    unsafe_allow_html=True,
)
st.markdown("---")

tab_results, tab_edit = st.tabs(["📊  Results", "✏️  Edit & Recalculate"])


# ══════════════════════════════════════════════════════════════════════════════
#  TAB 1 — RESULTS
# ══════════════════════════════════════════════════════════════════════════════
with tab_results:
    st.markdown("### Upload Composition Table")
    comp_file = st.file_uploader(
        "Upload your Composition_Table.xlsx",
        type=["xlsx"],
        key="comp_upload",
        help="Each sheet: COMPOSITION | NORMAL | UNIT columns. PPM rows auto-excluded.",
    )

    if comp_file:
        if st.button("▶  Calculate All Sheets"):
            with st.spinner("Loading database…"):
                try:
                    db = load_database()
                except Exception as e:
                    st.error(f"Failed to load database.xlsx: {e}")
                    st.stop()
            with st.spinner("Running calculations…"):
                try:
                    results, raw_mixtures = run_all_sheets(comp_file.read(), db, P_atm)
                    st.session_state.results      = results
                    st.session_state.raw_mixtures = raw_mixtures
                except Exception as e:
                    st.error(f"Calculation error: {e}")
                    st.stop()

    if st.session_state.results:
        results    = st.session_state.results
        ok_sheets  = {k: v for k, v in results.items() if "error" not in v}
        err_sheets = {k: v for k, v in results.items() if "error"     in v}

        c1, c2, c3 = st.columns(3)
        c1.metric("Sheets Processed", len(results))
        c2.metric("Successful",        len(ok_sheets))
        c3.metric("Errors / Skipped",  len(err_sheets))
        st.markdown("---")

        if ok_sheets:
            st.markdown("### Results Summary")
            rows = []
            for sname, res in ok_sheets.items():
                rows.append({
                    "Sheet":          sname.strip(),
                    "Components":     res["n"],
                    "Dew Pt (°C)":    round(res["dew_C"],    2),
                    "Dew Pt (K)":     round(res["dew_K"],    2),
                    "Bubble Pt (°C)": round(res["bubble_C"], 2),
                    "Bubble Pt (K)":  round(res["bubble_K"], 2),
                    "Note":           "⚠️" if res["not_found"] else "",
                })
            df_summary = pd.DataFrame(rows)
            st.dataframe(
                df_summary.style
                    .format({
                        "Dew Pt (°C)":    "{:.2f}",
                        "Dew Pt (K)":     "{:.2f}",
                        "Bubble Pt (°C)": "{:.2f}",
                        "Bubble Pt (K)":  "{:.2f}",
                    })
                    .set_properties(**{
                        "background-color": "#161b27",
                        "color":            "#e2e8f0",
                        "font-family":      "IBM Plex Mono, monospace",
                        "font-size":        "0.82rem",
                    })
                    .set_table_styles([{"selector": "th", "props": [
                        ("background-color", "#0f1117"), ("color", "#64748b"),
                        ("font-size", "0.72rem"), ("letter-spacing", "0.08em"),
                        ("text-transform", "uppercase"),
                        ("font-family", "IBM Plex Mono, monospace"),
                    ]}]),
                use_container_width=True,
                hide_index=True,
            )
            csv = df_summary.to_csv(index=False).encode("utf-8")
            st.download_button("⬇  Download Results as CSV",
                               data=csv, file_name="dew_bubble_results.csv",
                               mime="text/csv")

        if err_sheets:
            st.markdown("### ⚠️ Sheets with Errors")
            for sname, res in err_sheets.items():
                st.markdown(
                    f"<div class='result-card error'>"
                    f"<div style='font-family:IBM Plex Mono,monospace;font-size:0.85rem;"
                    f"color:#94a3b8;margin-bottom:0.3rem'>{sname.strip()}</div>"
                    f"<span style='color:#ef4444;font-size:0.85rem'>{res['error']}</span>"
                    f"</div>", unsafe_allow_html=True,
                )

        if ok_sheets:
            st.markdown("---")
            st.markdown("### Detailed Breakdown")
            for sname, res in ok_sheets.items():
                with st.expander(f"📋  {sname.strip()}  —  "
                                 f"Dew: {res['dew_C']:.2f} °C  |  "
                                 f"Bubble: {res['bubble_C']:.2f} °C"):
                    if res["not_found"]:
                        missing = ", ".join(nm for nm, _ in res["not_found"])
                        st.warning(f"**{len(res['not_found'])} component(s) not in database "
                                   f"and were excluded:** {missing}")
                    dc, bc = st.columns(2)
                    dc.metric("Dew Point",    f"{res['dew_C']:.2f} °C",
                              f"{res['dew_K']:.2f} K")
                    bc.metric("Bubble Point", f"{res['bubble_C']:.2f} °C",
                              f"{res['bubble_K']:.2f} K")
                    comp_df = pd.DataFrame(res["components"],
                                           columns=["Component", "Normalised Mol%"])
                    st.dataframe(
                        comp_df.style
                            .format({"Normalised Mol%": "{:.4f}"})
                            .bar(subset=["Normalised Mol%"], color="#0c4a6e", vmin=0),
                        use_container_width=True, hide_index=True,
                    )

    elif not comp_file:
        st.markdown("""
        <div style='background:#161b27;border:1px solid #2a3040;border-radius:8px;
                    padding:2rem 2.5rem;margin-top:1rem;'>
        <h3 style='color:#38bdf8;font-family:IBM Plex Mono,monospace;
                   font-size:1rem;margin-bottom:1rem;'>How to use</h3>
        <ol style='color:#94a3b8;font-size:0.88rem;line-height:2;'>
            <li>Set your <b>operating pressure</b> (kg/cm² gauge) in the sidebar</li>
            <li>Upload your <code>Composition_Table.xlsx</code> above</li>
            <li>Click <b>Calculate All Sheets</b></li>
            <li>Switch to <b>✏️ Edit &amp; Recalculate</b> tab to tweak any sheet's composition</li>
        </ol>
        <hr style='border-color:#2a3040;margin:1.2rem 0'>
        <p style='color:#475569;font-size:0.8rem;margin:0'>
            ℹ️ Only <code>MOL%</code> rows are used. PPM rows are auto-excluded.
        </p>
        </div>
        """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  TAB 2 — EDIT & RECALCULATE
# ══════════════════════════════════════════════════════════════════════════════
with tab_edit:
    if not st.session_state.results:
        st.info("👈  Run a calculation first in the **📊 Results** tab, then come back here.")
    else:
        results      = st.session_state.results
        raw_mixtures = st.session_state.raw_mixtures

        # Only successfully calculated sheets are available to edit
        ok_sheet_names = [k for k, v in results.items() if "error" not in v]

        st.markdown("### Select a Sheet to Edit")
        selected_sheet = st.selectbox(
            "Sheet",
            options=ok_sheet_names,
            format_func=lambda s: s.strip(),
            key="edit_sheet_select",
        )

        if selected_sheet:
            original_result  = results[selected_sheet]
            original_mixture = raw_mixtures[selected_sheet]

            # Original result for reference
            st.markdown("---")
            st.markdown("#### Original Results")
            oc1, oc2 = st.columns(2)
            oc1.metric("Dew Point",    f"{original_result['dew_C']:.2f} °C",
                       f"{original_result['dew_K']:.2f} K")
            oc2.metric("Bubble Point", f"{original_result['bubble_C']:.2f} °C",
                       f"{original_result['bubble_K']:.2f} K")

            st.markdown("---")
            st.markdown("#### Edit Composition")
            st.caption(
                "Change mol% values, rename components, add new rows or delete existing ones. "
                "Values are re-normalised to 100% automatically before recalculating."
            )

            # Build editable dataframe from original mixture
            edit_df = pd.DataFrame(original_mixture, columns=["Component", "Mol%"])

            edited_df = st.data_editor(
                edit_df,
                num_rows="dynamic",
                use_container_width=True,
                column_config={
                    "Component": st.column_config.TextColumn(
                        "Component",
                        help="Must match a name in the database exactly",
                        width="large",
                    ),
                    "Mol%": st.column_config.NumberColumn(
                        "Mol%",
                        help="Mole percentage — will be normalised to 100% automatically",
                        min_value=0.0,
                        format="%.4f",
                        width="medium",
                    ),
                },
                key=f"editor_{selected_sheet}",
            )

            if st.button("🔄  Recalculate", key="recalc_btn"):

                # Parse edited dataframe into mixture list
                new_mixture = []
                for _, row in edited_df.iterrows():
                    name = str(row["Component"]).strip().upper() \
                           if pd.notna(row["Component"]) else ""
                    val  = row["Mol%"]
                    if not name or not isinstance(val, (int, float)) \
                            or pd.isna(val) or val <= 0:
                        continue
                    new_mixture.append((name, float(val)))

                if not new_mixture:
                    st.error("No valid components found. "
                             "Make sure component names and mol% values are filled in.")
                else:
                    with st.spinner("Recalculating…"):
                        try:
                            db         = load_database()
                            new_result = process_mixture(new_mixture, db, P_atm)
                        except Exception as e:
                            st.error(f"Error: {e}")
                            st.stop()

                    st.markdown("---")

                    if "error" in new_result:
                        st.error(f"Calculation failed: {new_result['error']}")
                    else:
                        if new_result["not_found"]:
                            missing = ", ".join(nm for nm, _ in new_result["not_found"])
                            st.warning(
                                f"**{len(new_result['not_found'])} component(s) not found "
                                f"in database and were excluded:** {missing}"
                            )

                        st.markdown("#### Recalculated Results")
                        nc1, nc2 = st.columns(2)
                        dew_delta    = new_result["dew_C"]    - original_result["dew_C"]
                        bubble_delta = new_result["bubble_C"] - original_result["bubble_C"]

                        nc1.metric(
                            "New Dew Point",
                            f"{new_result['dew_C']:.2f} °C",
                            delta=f"{dew_delta:+.2f} °C vs original",
                        )
                        nc2.metric(
                            "New Bubble Point",
                            f"{new_result['bubble_C']:.2f} °C",
                            delta=f"{bubble_delta:+.2f} °C vs original",
                        )

                        # Side-by-side comparison
                        st.markdown("#### Comparison")
                        cmp_df = pd.DataFrame({
                            "":               ["Dew Point (°C)", "Dew Point (K)",
                                               "Bubble Point (°C)", "Bubble Point (K)"],
                            "Original":       [round(original_result["dew_C"],    2),
                                               round(original_result["dew_K"],    2),
                                               round(original_result["bubble_C"], 2),
                                               round(original_result["bubble_K"], 2)],
                            "Recalculated":   [round(new_result["dew_C"],    2),
                                               round(new_result["dew_K"],    2),
                                               round(new_result["bubble_C"], 2),
                                               round(new_result["bubble_K"], 2)],
                            "Δ Change":       [round(dew_delta,                              2),
                                               round(new_result["dew_K"]    - original_result["dew_K"],    2),
                                               round(bubble_delta,                           2),
                                               round(new_result["bubble_K"] - original_result["bubble_K"], 2)],
                        })

                        def colour_delta(val):
                            if isinstance(val, (int, float)):
                                if val > 0:  return "color: #4ade80"
                                if val < 0:  return "color: #f87171"
                            return ""

                        st.dataframe(
                            cmp_df.style
                                .applymap(colour_delta, subset=["Δ Change"])
                                .set_properties(**{
                                    "background-color": "#161b27",
                                    "color":            "#e2e8f0",
                                    "font-family":      "IBM Plex Mono, monospace",
                                    "font-size":        "0.84rem",
                                })
                                .set_table_styles([{"selector": "th", "props": [
                                    ("background-color", "#0f1117"),
                                    ("color",            "#64748b"),
                                    ("font-family",      "IBM Plex Mono, monospace"),
                                    ("font-size",        "0.72rem"),
                                    ("text-transform",   "uppercase"),
                                ]}]),
                            use_container_width=True,
                            hide_index=True,
                        )

                        # Normalised composition that was actually used
                        st.markdown("#### Composition Used (normalised)")
                        used_df = pd.DataFrame(
                            new_result["components"],
                            columns=["Component", "Normalised Mol%"],
                        )
                        st.dataframe(
                            used_df.style
                                .format({"Normalised Mol%": "{:.4f}"})
                                .bar(subset=["Normalised Mol%"], color="#0c4a6e", vmin=0),
                            use_container_width=True,
                            hide_index=True,
                        )