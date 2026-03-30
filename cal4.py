import math
import pandas as pd
from openpyxl import load_workbook

# ==============================================================================
# CONFIGURATION — change these as needed
# ==============================================================================
PRESSURE_KGCM2_GAUGE  = 1.5          # kg/cm² gauge
DATABASE_FILE         = "database.xlsx"
COMPOSITION_FILE      = "Composition_Table.xlsx"

# Units in the composition table that count as MOLE (not PPM, not skipped).
# Anything containing "PPM" in the unit cell is excluded automatically.
MOL_UNIT_KEYWORDS = ["MOL%", "MOL %", "MOLE%", "MOLE %"]
# ==============================================================================


# ── Helpers ────────────────────────────────────────────────────────────────────

def is_mol_unit(unit_str):
    """Return True if the unit string represents a mole-fraction / mole-percent."""
    if not unit_str:
        return False
    u = str(unit_str).strip().upper()
    # Exclude anything with PPM in it
    if "PPM" in u:
        return False
    return any(kw in u for kw in [k.upper() for k in MOL_UNIT_KEYWORDS])


def convert_pressure(P_gauge_kgcm2):
    """kg/cm² gauge  →  ATM absolute."""
    P_abs_kgcm2 = P_gauge_kgcm2 + 1.0332   # add 1 atm in kg/cm²
    return P_abs_kgcm2 / 1.0332             # convert to ATM


def wilson_k(Pc_atm, P_atm, omega, Tc_K, T_K):
    """Wilson correlation K-value (identical to VBA formula)."""
    return (Pc_atm / P_atm) * math.exp(5.37 * (1.0 + omega) * (1.0 - Tc_K / T_K))


# ── Database loader ────────────────────────────────────────────────────────────

def load_database(file_path=DATABASE_FILE):
    """
    Skip the 7-row header block and assign column names.
    Column order matches the database header exactly.
    """
    df = pd.read_excel(file_path, header=None, skiprows=7)
    df.columns = [
        'NUMBER', 'COMPONENT', 'MOLE_WT', 'TFP', 'TB',
        'TC',     # critical temperature, K
        'PC',     # critical pressure,    ATM
        'VC', 'ZC',
        'OMEGA',  # acentric factor
        'LIQDEN', 'TDEN', 'DIPM',
        'CP_A', 'CP_B', 'CP_C', 'CP_D',
        'LV_B', 'LV_C',
        'DELHG', 'DELGF',
        'ANT_A', 'ANT_B', 'ANT_C', 'TMX', 'TMN',
        'HAR_A', 'HAR_B', 'HAR_C', 'HAR_D',
        'HV',
    ]
    df = df.dropna(subset=['COMPONENT'])
    df['COMPONENT'] = df['COMPONENT'].astype(str).str.strip().str.upper()
    return df


# ── Composition reader ─────────────────────────────────────────────────────────

def read_composition_sheet(ws):
    """
    Read one sheet from the composition workbook.
    Returns a list of (name, mole_percent) for MOL% rows only.
    Skips:  PPM rows, blank rows, the header row (row 0), zero-value rows.
    """
    mixture = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:          # header row: COMPOSITION / NORMAL / UNIT
            continue

        name  = row[0]
        value = row[1]
        unit  = row[2] if len(row) > 2 else None

        # Skip blank rows
        if name is None or value is None:
            continue

        # Skip non-numeric values (e.g. "TRACE")
        try:
            value = float(value)
        except (TypeError, ValueError):
            continue

        # Skip zero-value components (they contribute nothing)
        if value == 0:
            continue

        # Only keep MOL% rows
        if not is_mol_unit(unit):
            continue

        mixture.append((str(name).strip().upper(), value))

    return mixture


# ── Core calculation ───────────────────────────────────────────────────────────

def calculate_dew_bubble(P_atm, component_data, mole_fractions):
    """
    Replicates VBA loop exactly:
      T_n = 0.25 K, step 0.25 K, 2500 iterations  (max 625 K)
      TOL = 1e-300
    """
    TOL = 1e-300
    T_n = 0.25
    step = 0.25
    iterations = 2500

    best_dew    = (float('inf'), None)
    best_bubble = (float('inf'), None)

    for _ in range(iterations):
        dew_sum    = 0.0
        bubble_sum = 0.0

        for z_i, props in zip(mole_fractions, component_data):
            k = wilson_k(props['PC'], P_atm, props['OMEGA'], props['TC'], T_n)

            if k >= TOL:          # dew:    y / k
                dew_sum += z_i / k
            if k != 0:            # bubble: x * k
                bubble_sum += z_i * k

        d_diff = abs(dew_sum    - 1)
        b_diff = abs(bubble_sum - 1)

        if d_diff < best_dew[0]:
            best_dew = (d_diff, T_n)
        if b_diff < best_bubble[0]:
            best_bubble = (b_diff, T_n)

        T_n += step

    return best_dew[1], best_bubble[1]


# ── Per-sheet runner ───────────────────────────────────────────────────────────

def process_sheet(sheet_name, mixture, db, P_atm):
    """
    Look up each component in the database, normalise fractions,
    run the Wilson calculation.
    Returns a dict with results (or an error message).
    """
    if not mixture:
        return {"error": "No MOL% components found (all PPM or all zero)"}

    component_data  = []
    mole_percents   = []
    not_found       = []

    for name, pct in mixture:
        row = db[db['COMPONENT'] == name]
        if row.empty:
            # Try stripping trailing spaces / common aliases
            partial = db[db['COMPONENT'].str.contains(name, na=False, regex=False)]
            hint = partial['COMPONENT'].tolist()[:3] if not partial.empty else []
            not_found.append((name, hint))
            continue
        component_data.append(row.iloc[0])
        mole_percents.append(pct)

    if not_found:
        msgs = []
        for nm, hints in not_found:
            h = f"  → closest: {hints}" if hints else "  → not in database"
            msgs.append(f"    '{nm}'{h}")
        warning = "Some components skipped (not in database):\n" + "\n".join(msgs)
    else:
        warning = None

    if not component_data:
        return {"error": "No components could be matched in the database",
                "warning": warning}

    # Normalise to fractions (values are in mol %)
    total = sum(mole_percents)
    if total == 0:
        return {"error": "All mole percents are zero"}

    mole_fractions = [p / total for p in mole_percents]

    dew_K, bubble_K = calculate_dew_bubble(P_atm, component_data, mole_fractions)

    return {
        "n_components"  : len(component_data),
        "components"    : [(c['COMPONENT'], round(f * 100, 4))
                           for c, f in zip(component_data, mole_fractions)],
        "dew_K"         : dew_K,
        "dew_C"         : dew_K - 273.15,
        "bubble_K"      : bubble_K,
        "bubble_C"      : bubble_K - 273.15,
        "warning"       : warning,
    }


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print("=" * 70)
    print("  DEW & BUBBLE POINT CALCULATOR  —  Wilson Correlation")
    print("=" * 70)
    print(f"  Pressure  : {PRESSURE_KGCM2_GAUGE} kg/cm² gauge")

    P_atm = convert_pressure(PRESSURE_KGCM2_GAUGE)
    print(f"            = {P_atm:.4f} ATM absolute")
    print(f"  Database  : {DATABASE_FILE}")
    print(f"  Compos.   : {COMPOSITION_FILE}")
    print("=" * 70)

    # Load database
    db = load_database(DATABASE_FILE)

    # Open composition workbook
    wb = load_workbook(COMPOSITION_FILE, read_only=True, data_only=True)

    all_results = {}

    for sheet_name in wb.sheetnames:
        ws      = wb[sheet_name]
        mixture = read_composition_sheet(ws)
        result  = process_sheet(sheet_name, mixture, db, P_atm)
        all_results[sheet_name] = result

    wb.close()

    # ── Print results ──────────────────────────────────────────────────────────
    print(f"\n{'Sheet':<35} {'Dew Pt (°C)':>12} {'Bubble Pt (°C)':>15}  {'Notes'}")
    print("-" * 90)

    for sheet_name, res in all_results.items():
        short = sheet_name.strip()[:34]

        if "error" in res:
            print(f"{short:<35} {'—':>12} {'—':>15}  ERROR: {res['error']}")
        else:
            dew_str    = f"{res['dew_C']:>10.2f} °C"
            bubble_str = f"{res['bubble_C']:>13.2f} °C"
            note       = f"({res['n_components']} components)"
            if res['warning']:
                note += "  ⚠ " + res['warning'].split('\n')[0]
            print(f"{short:<35} {dew_str} {bubble_str}  {note}")

    # ── Detailed breakdown ─────────────────────────────────────────────────────
    print("\n\n" + "=" * 70)
    print("  DETAILED RESULTS")
    print("=" * 70)

    for sheet_name, res in all_results.items():
        print(f"\n▶  Sheet: {sheet_name.strip()}")
        if "error" in res:
            print(f"   ERROR: {res['error']}")
            if res.get('warning'):
                print(f"   {res['warning']}")
            continue

        if res['warning']:
            print(f"   ⚠  {res['warning']}")

        print(f"   Components used ({res['n_components']}):")
        for cname, cpct in res['components']:
            print(f"      {cname:<30}  {cpct:>8.4f} mol%")

        print(f"\n   Dew    Point : {res['dew_K']:.2f} K  =  {res['dew_C']:.2f} °C")
        print(f"   Bubble Point : {res['bubble_K']:.2f} K  =  {res['bubble_C']:.2f} °C")


if __name__ == "__main__":
    main()